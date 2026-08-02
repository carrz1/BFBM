"""Merge existing 55 'Done' rows from HRB_System_Performance_Audit_noggin.xlsx
with the 37 newly-audited pending slots (computed in-browser via fetch, not
locally cached as raw TSV) into one final noggin (account 1) audit workbook.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

SRC = Path(__file__).parent / "HRB_System_Performance_Audit_noggin.xlsx"
OUT = Path(__file__).parent / "HRB_System_Performance_Audit_noggin_FINAL.xlsx"

# newly computed slots (2026-07-31), via in-browser fetch + JS port of the
# odds-band-parse + saved+1-filter + BF-commission-formula pipeline validated
# against the Python compute() in build_noggin2_audit.py (cross-checked on
# slot 58: JS gave Bets 160/Wins 1/P&L(SP) -119/P&L(BF) -69.70, Python
# independently gave Bets 160/Wins 1/P&L(SP) -119.0/P&L(BF) -69.70 - match).
NEW = [
    {"Slot":58,"Name":"8YO 9YO 18.01 - 1000.00","Saved":"2021-04-13","OddsBand":"18.01 - 1000","Note":"","Bets":160,"Wins":1,"Win%":0.63,"P/L(SP)":-119,"Races":154,"Race%":0.65,"ROI(SP)":-74.38,"P/L(BF)":-69.7,"ROI(BF)":-43.56,"ByYear":{2021:-26,2022:-28,2023:-30,2024:-28,2025:-24,2026:66.3}},
    {"Slot":60,"Name":"1m2f - 2m.5f hood blinkers","Saved":"2019-01-31","OddsBand":"none","Note":"","Bets":1890,"Wins":246,"Win%":13.02,"P/L(SP)":-202.08,"Races":1687,"Race%":14.58,"ROI(SP)":-10.69,"P/L(BF)":135.84,"ROI(BF)":7.19,"ByYear":{2019:29.08,2020:65.54,2021:7.08,2022:-32.17,2023:62.69,2024:-27.94,2025:44.8,2026:-13.25}},
    {"Slot":61,"Name":"ss DoncSTR 4YOs","Saved":"2023-09-15","OddsBand":"none","Note":"","Bets":270,"Wins":19,"Win%":7.04,"P/L(SP)":-61.92,"Races":87,"Race%":21.84,"ROI(SP)":-22.93,"P/L(BF)":-16.97,"ROI(BF)":-6.28,"ByYear":{2023:-8,2024:27.22,2025:0.41,2026:-36.6}},
    {"Slot":62,"Name":"John P McManus max30.00","Saved":"2026-02-22","OddsBand":"none","Note":"","Bets":125,"Wins":20,"Win%":16,"P/L(SP)":-31.64,"Races":107,"Race%":18.69,"ROI(SP)":-25.31,"P/L(BF)":-15.98,"ROI(BF)":-12.78,"ByYear":{2026:-15.98}},
    {"Slot":63,"Name":"5YO 6f AUTUMN FMs","Saved":"2022-01-27","OddsBand":"none","Note":"","Bets":48,"Wins":5,"Win%":10.42,"P/L(SP)":13.63,"Races":44,"Race%":11.36,"ROI(SP)":28.4,"P/L(BF)":28.05,"ROI(BF)":58.44,"ByYear":{2022:-6.34,2023:44.18,2024:-2.72,2025:-7.06}},
    {"Slot":64,"Name":"FIRST RUN OF SEASON","Saved":"2021-04-25","OddsBand":"none","Note":"","Bets":107,"Wins":13,"Win%":12.15,"P/L(SP)":64.75,"Races":103,"Race%":12.62,"ROI(SP)":60.51,"P/L(BF)":152.32,"ROI(BF)":142.35,"ByYear":{2021:-13,2022:6.44,2023:126.78,2024:6.36,2025:-10.72,2026:36.46}},
    {"Slot":65,"Name":"5YO 6f AUTUMN","Saved":"2012-09-13","OddsBand":"none","Note":"","Bets":500,"Wins":63,"Win%":12.6,"P/L(SP)":55.19,"Races":412,"Race%":15.29,"ROI(SP)":11.04,"P/L(BF)":280.26,"ROI(BF)":56.05,"ByYear":{2012:-10.72,2013:27.58,2014:3.9,2015:93.71,2016:37.76,2017:86.88,2018:93.36,2019:17.65,2020:11.78,2021:10.05,2022:-26.43,2023:-28.33,2024:-17.33,2025:-19.62}},
    {"Slot":66,"Name":"SS AW WON HIGHER CLASS 8.01-1000.00","Saved":"2025-02-02","OddsBand":"8.01 - 1000","Note":"","Bets":127,"Wins":13,"Win%":10.24,"P/L(SP)":190.5,"Races":119,"Race%":10.92,"ROI(SP)":150,"P/L(BF)":270.62,"ROI(BF)":213.08,"ByYear":{2025:178.31,2026:92.3}},
    {"Slot":67,"Name":"SS NH C WINNER 30.01 - 1000.00","Saved":"2025-10-02","OddsBand":"30.01 - 1000","Note":"","Bets":138,"Wins":4,"Win%":2.9,"P/L(SP)":-6,"Races":118,"Race%":3.39,"ROI(SP)":-4.35,"P/L(BF)":80.62,"ROI(BF)":58.42,"ByYear":{2025:-41,2026:121.62}},
    {"Slot":68,"Name":"SSGB 4YOs TURF 12.01 - 1000.00","Saved":"2025-10-20","OddsBand":"12.01 - 1000","Note":"","Bets":355,"Wins":9,"Win%":2.54,"P/L(SP)":-220,"Races":263,"Race%":3.42,"ROI(SP)":-61.97,"P/L(BF)":-189.69,"ROI(BF)":-53.43,"ByYear":{2025:-0.85,2026:-188.84}},
    {"Slot":69,"Name":"ROYAL ASCOT GB 31 plus DSLR","Saved":"2021-06-18","OddsBand":"none","Note":"","Bets":133,"Wins":4,"Win%":3.01,"P/L(SP)":-61.5,"Races":75,"Race%":5.33,"ROI(SP)":-46.24,"P/L(BF)":-18.03,"ROI(BF)":-13.56,"ByYear":{2021:-6,2022:49.3,2023:14.48,2024:-21,2025:-34,2026:-20.82}},
    {"Slot":70,"Name":"SS MM NAAS 3.01 - 100.00 MM","Saved":"2024-02-26","OddsBand":"3.01 - 100","Note":"","Bets":33,"Wins":4,"Win%":12.12,"P/L(SP)":-7.75,"Races":24,"Race%":16.67,"ROI(SP)":-23.48,"P/L(BF)":-4.12,"ROI(BF)":-12.48,"ByYear":{2024:-11.67,2025:8.63,2026:-1.08}},
    {"Slot":71,"Name":"ROYAL ASCOT GB 31 plus DSLR v2","Saved":"2021-06-18","OddsBand":"none","Note":"","Bets":156,"Wins":4,"Win%":2.56,"P/L(SP)":-84.5,"Races":84,"Race%":4.76,"ROI(SP)":-54.17,"P/L(BF)":-41.03,"ROI(BF)":-26.3,"ByYear":{2021:-6,2022:49.3,2023:7.48,2024:-32,2025:-39,2026:-20.82}},
    {"Slot":73,"Name":"SS Williams I","Saved":"2023-05-13","OddsBand":"none","Note":"","Bets":90,"Wins":6,"Win%":6.67,"P/L(SP)":-10.25,"Races":88,"Race%":6.82,"ROI(SP)":-11.39,"P/L(BF)":9.79,"ROI(BF)":10.88,"ByYear":{2023:-21.59,2024:16.38,2025:27,2026:-12}},
    {"Slot":74,"Name":"EBOR GB 8.01 PLUS","Saved":"2021-08-18","OddsBand":"none","Note":"","Bets":122,"Wins":8,"Win%":6.56,"P/L(SP)":-19.62,"Races":66,"Race%":12.12,"ROI(SP)":-16.08,"P/L(BF)":1.56,"ROI(BF)":1.28,"ByYear":{2021:15.21,2022:-15.1,2023:-17.74,2024:26.67,2025:-7.48}},
    {"Slot":75,"Name":"MM NH 0-7 runs HDGR 4.01 - 1000","Saved":"2024-01-07","OddsBand":"4.01 - 1000","Note":"skipped '0-7' (value <1.0, not odds)","Bets":4131,"Wins":367,"Win%":8.88,"P/L(SP)":-506.99,"Races":3237,"Race%":11.34,"ROI(SP)":-12.27,"P/L(BF)":1119.51,"ROI(BF)":27.1,"ByYear":{2024:-58.75,2025:1142.22,2026:36.05}},
    {"Slot":78,"Name":"ROYAL ASCOT GB 12 - 20 DSLR","Saved":"2021-06-18","OddsBand":"none","Note":"skipped '12 - 20' (followed by 'dslr', not odds)","Bets":37,"Wins":3,"Win%":8.11,"P/L(SP)":37,"Races":29,"Race%":10.34,"ROI(SP)":100,"P/L(BF)":92.36,"ROI(BF)":249.62,"ByYear":{2021:-1,2022:108.33,2023:-4,2024:-6,2025:1.03,2026:-6}},
    {"Slot":79,"Name":"NH HDGRDiff 6.01 - 1000.00 CHP and NONE","Saved":"2024-07-09","OddsBand":"6.01 - 1000","Note":"","Bets":1628,"Wins":98,"Win%":6.02,"P/L(SP)":-408,"Races":1349,"Race%":7.26,"ROI(SP)":-25.06,"P/L(BF)":82.8,"ROI(BF)":5.09,"ByYear":{2024:-1.19,2025:231.27,2026:-147.28}},
    {"Slot":80,"Name":"ssDonc 5f-6.5f 4.51 - 8.00","Saved":"2023-09-16","OddsBand":"4.51 - 8","Note":"","Bets":17,"Wins":4,"Win%":23.53,"P/L(SP)":4.5,"Races":16,"Race%":25,"ROI(SP)":26.47,"P/L(BF)":7.13,"ROI(BF)":41.94,"ByYear":{2024:0.56,2025:0.32,2026:6.25}},
    {"Slot":82,"Name":"MM","Saved":"2025-05-20","OddsBand":"none","Note":"CAPPED AT 10,000-ROW DOWNLOAD LIMIT - true post-saved-date bet count may be higher, review manually","Bets":10000,"Wins":1795,"Win%":17.95,"P/L(SP)":-1634.25,"Races":4602,"Race%":38.98,"ROI(SP)":-16.34,"P/L(BF)":-306.78,"ROI(BF)":-3.07,"ByYear":{2026:-306.78}},
    {"Slot":83,"Name":"SS SouthwellAW 8f - 12f","Saved":"2024-12-19","OddsBand":"none","Note":"","Bets":95,"Wins":10,"Win%":10.53,"P/L(SP)":-41,"Races":75,"Race%":13.33,"ROI(SP)":-43.16,"P/L(BF)":-33.56,"ROI(BF)":-35.32,"ByYear":{2024:3.9,2025:-13.35,2026:-24.11}},
    {"Slot":84,"Name":"ss HDGRdiff IRE horses SPRING all BSPs","Saved":"2023-09-23","OddsBand":"none","Note":"name says all BSPs = unrestricted","Bets":682,"Wins":46,"Win%":6.74,"P/L(SP)":-108.79,"Races":622,"Race%":7.4,"ROI(SP)":-15.95,"P/L(BF)":60.18,"ROI(BF)":8.82,"ByYear":{2024:52.07,2025:37.23,2026:-29.13}},
    {"Slot":85,"Name":"MSS MM LUDLOW MM 7.01 - 100.00","Saved":"2024-02-21","OddsBand":"7.01 - 100","Note":"","Bets":45,"Wins":4,"Win%":8.89,"P/L(SP)":10,"Races":41,"Race%":9.76,"ROI(SP)":22.22,"P/L(BF)":16,"ROI(BF)":35.56,"ByYear":{2024:3.67,2025:10.51,2026:1.82}},
    {"Slot":87,"Name":"MM NH CARLISLE 2.01 - 100.00","Saved":"2024-02-19","OddsBand":"2.01 - 100","Note":"","Bets":159,"Wins":36,"Win%":22.64,"P/L(SP)":36.47,"Races":105,"Race%":34.29,"ROI(SP)":22.94,"P/L(BF)":63.67,"ROI(BF)":40.04,"ByYear":{2024:49,2025:10.93,2026:3.73}},
    {"Slot":88,"Name":"MM NH FONTWELL MM 2.51 - 300.00","Saved":"2024-01-28","OddsBand":"2.51 - 300","Note":"","Bets":46,"Wins":7,"Win%":15.22,"P/L(SP)":11.25,"Races":45,"Race%":15.56,"ROI(SP)":24.46,"P/L(BF)":23.31,"ROI(BF)":50.67,"ByYear":{2024:-1.02,2025:29.59,2026:-5.26}},
    {"Slot":89,"Name":"SS HDGRdiff 80.01 - 250.00","Saved":"2025-03-09","OddsBand":"80.01 - 250","Note":"","Bets":149,"Wins":2,"Win%":1.34,"P/L(SP)":44,"Races":145,"Race%":1.38,"ROI(SP)":29.53,"P/L(BF)":145.6,"ROI(BF)":97.72,"ByYear":{2025:208.6,2026:-63}},
    {"Slot":90,"Name":"MM AW SOUTHWELL 7.01 - 50.00 MM","Saved":"2024-01-31","OddsBand":"7.01 - 50","Note":"","Bets":456,"Wins":43,"Win%":9.43,"P/L(SP)":-74,"Races":332,"Race%":12.95,"ROI(SP)":-16.23,"P/L(BF)":-14.79,"ROI(BF)":-3.24,"ByYear":{2024:12.6,2025:-11.04,2026:-16.35}},
    {"Slot":91,"Name":"STRAIGHT HDGR 4YO","Saved":"2018-05-21","OddsBand":"none","Note":"","Bets":88,"Wins":7,"Win%":7.95,"P/L(SP)":48.5,"Races":75,"Race%":9.33,"ROI(SP)":55.11,"P/L(BF)":90.89,"ROI(BF)":103.29,"ByYear":{2018:1.72,2019:0.88,2020:40.41,2021:-12,2022:21.09,2023:23.15,2024:-12,2025:29.65,2026:-2}},
    {"Slot":92,"Name":"SS 4YO STRAIGHTS Turf All BSPs","Saved":"2023-01-13","OddsBand":"none","Note":"name says All BSPs = unrestricted","Bets":123,"Wins":14,"Win%":11.38,"P/L(SP)":-26.34,"Races":83,"Race%":16.87,"ROI(SP)":-21.41,"P/L(BF)":-8.68,"ROI(BF)":-7.06,"ByYear":{2023:-14.59,2024:2.12,2025:6.89,2026:-3.09}},
    {"Slot":93,"Name":"MM 9.01 - 200 NH and AW","Saved":"2025-11-01","OddsBand":"9.01 - 200","Note":"","Bets":1367,"Wins":83,"Win%":6.07,"P/L(SP)":-327.5,"Races":1057,"Race%":7.85,"ROI(SP)":-23.96,"P/L(BF)":-71.89,"ROI(BF)":-5.26,"ByYear":{2025:4.73,2026:-76.62}},
    {"Slot":94,"Name":"IRISH FLAT FM 10.01 - 200.00","Saved":"2025-05-20","OddsBand":"10.01 - 200","Note":"","Bets":2419,"Wins":116,"Win%":4.8,"P/L(SP)":-465.5,"Races":782,"Race%":14.83,"ROI(SP)":-19.24,"P/L(BF)":513.81,"ROI(BF)":21.24,"ByYear":{2025:309.03,2026:204.78}},
    {"Slot":95,"Name":"SS STRTs DRAW Q1 Q4 10.01 - 1000.00","Saved":"2023-10-08","OddsBand":"10.01 - 1000","Note":"","Bets":521,"Wins":28,"Win%":5.37,"P/L(SP)":-10.5,"Races":296,"Race%":9.46,"ROI(SP)":-2.02,"P/L(BF)":141.05,"ROI(BF)":27.07,"ByYear":{2023:-16,2024:59.55,2025:97.01,2026:0.49}},
    {"Slot":96,"Name":"IRISH FLAT FM v1 11.01 - 200.00","Saved":"2026-05-17","OddsBand":"11.01 - 200","Note":"","Bets":515,"Wins":22,"Win%":4.27,"P/L(SP)":-186.5,"Races":187,"Race%":11.76,"ROI(SP)":-36.21,"P/L(BF)":-19.08,"ROI(BF)":-3.71,"ByYear":{2026:-19.08}},
    {"Slot":97,"Name":"MM NH 10.01 - 80.00","Saved":"2025-05-20","OddsBand":"10.01 - 80","Note":"","Bets":4181,"Wins":203,"Win%":4.86,"P/L(SP)":-1284.5,"Races":1263,"Race%":16.07,"ROI(SP)":-30.72,"P/L(BF)":-34.88,"ROI(BF)":-0.83,"ByYear":{2025:-336.89,2026:302.01}},
    {"Slot":98,"Name":"MM IRISH FLAT 6.01 -200.00 HDGR max 40.00","Saved":"2025-05-20","OddsBand":"6.01 - 200","Note":"","Bets":1037,"Wins":58,"Win%":5.59,"P/L(SP)":-338.67,"Races":670,"Race%":8.66,"ROI(SP)":-32.66,"P/L(BF)":4.32,"ROI(BF)":0.42,"ByYear":{2025:187.6,2026:-183.28}},
    {"Slot":99,"Name":"SS HDGRdiff STRTs 7.01 - 500.00","Saved":"2024-07-27","OddsBand":"7.01 - 500","Note":"","Bets":380,"Wins":23,"Win%":6.05,"P/L(SP)":44,"Races":300,"Race%":7.67,"ROI(SP)":11.58,"P/L(BF)":172.06,"ROI(BF)":45.28,"ByYear":{2024:-18.47,2025:117.11,2026:73.42}},
    {"Slot":100,"Name":"DD CLASS DOWN v1","Saved":"2021-08-20","OddsBand":"none","Note":"","Bets":537,"Wins":61,"Win%":11.36,"P/L(SP)":-53.76,"Races":504,"Race%":12.1,"ROI(SP)":-10.01,"P/L(BF)":14.18,"ROI(BF)":2.64,"ByYear":{2021:-11.78,2022:147.34,2023:-0.81,2024:-49.48,2025:-61.62,2026:-9.47}},
]

wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_src = wb_src.active
headers_src = [c.value for c in ws_src[1]]

existing_rows = []
for row in ws_src.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers_src, row))
    if d.get("Slot") is None:
        continue
    by_year = {}
    text = d.get("P/L by Year (BF)")
    if text and isinstance(text, str):
        for part in text.split(";"):
            part = part.strip()
            if not part:
                continue
            y, v = part.split(":", 1)
            by_year[int(y.strip())] = float(v.strip())
    elif text:
        print(f"WARNING: slot {d['Slot']} ({d['System Name']}) has a corrupted "
              f"'P/L by Year (BF)' cell (Excel auto-formatted it as {type(text).__name__}: {text!r}) "
              f"from the original account-1 workbook - by-year breakdown skipped for this row, "
              f"totals unaffected.")
    existing_rows.append({
        "Slot": d["Slot"], "Name": d["System Name"],
        "Saved": d["Saved Date"].strftime("%Y-%m-%d") if hasattr(d["Saved Date"], "strftime") else d["Saved Date"],
        "OddsBand": d.get("Odds Band (BFSP) Applied") or "none",
        "Note": d.get("Odds Parsing Note") or "",
        "Bets": d["Bets"], "Wins": d["Wins"], "Win%": d["Win%"],
        "P/L(SP)": d["P/L(SP)"], "Races": d["Races"], "Race%": d["Race%"],
        "ROI(SP)": d["ROI(SP)%"], "P/L(BF)": d["P/L(BF)"], "ROI(BF)": d["ROI(BF)%"],
        "ByYear": by_year,
    })

all_rows = existing_rows + NEW
all_rows.sort(key=lambda r: r["Slot"])

years = sorted({y for r in all_rows for y in r["ByYear"].keys()})

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "noggin"
headers = ["Slot", "Name", "Saved", "Odds Band", "Odds Parsing Note", "Bets",
           "Wins", "Win%", "P/L(SP)", "Races", "Race%", "ROI(SP)", "P/L(BF)",
           "ROI(BF)"] + [f"P/L(BF) {y}" for y in years]
ws.append(headers)
for r in all_rows:
    row = [r["Slot"], r["Name"], r["Saved"], r["OddsBand"], r["Note"],
           r["Bets"], r["Wins"], r["Win%"], r["P/L(SP)"], r["Races"],
           r["Race%"], r["ROI(SP)"], r["P/L(BF)"], r["ROI(BF)"]]
    for y in years:
        row.append(r["ByYear"].get(y, ""))
    ws.append(row)

for i, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

wb.save(OUT)
print(f"Wrote {OUT} with {len(all_rows)} systems ({len(existing_rows)} pre-existing + {len(NEW)} newly audited).")
