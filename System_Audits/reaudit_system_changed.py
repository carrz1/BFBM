"""Re-audit the 17 systems flagged 'system-changed' by cross_account_overlap.py.

These are NOT the ordinary daily-qualifier drift that explains most of the
raw-vs-workbook mismatches (see PROJECT.md / CHANGELOG 2026-08-01) - the
evidence (mostly uncapped slots showing FEWER bets than the workbook, which
accumulation cannot cause) points to the underlying system criteria having
actually changed since the workbook was built: edited criteria, or dynamic
(form/ratings-relative) criteria re-evaluating against newer data.

16 are noggin slots 1-57 (the rows build_noggin_final.py carried over
verbatim from the older account-1 workbook and never recomputed), plus
noggin3 slot 67.

This recomputes all 17 from the currently-cached raw TSVs (downloaded
2026-08-01, i.e. current as of the re-audit), reusing:
  - the Odds Band already validated and stored in each workbook (not
    re-parsed from the name - no need to re-derive something already fixed)
  - the exact same saved+1-date filter / BF commission formula as every
    other script in this project (build_noggin2_audit.compute)

then writes the updated Bets/Wins/.../P/L(BF) by-year cells back into the
two live workbooks in place, leaving every other row untouched.
"""
import re
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

BASE = Path(__file__).parent

TARGETS = {
    "noggin": {
        "wb": "HRB_System_Performance_Audit_noggin_FINAL.xlsx",
        "raw": "noggin_raw",
        "slots": [1, 6, 27, 30, 33, 34, 37, 40, 43, 44, 49, 51, 53, 54, 55, 57],
    },
    "noggin3": {
        "wb": "HRB_System_Performance_Audit_noggin3.xlsx",
        "raw": "noggin3_raw",
        "slots": [67],
    },
}

BAND_RE = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)\s*$")


def parse_band(cell):
    if cell is None or (isinstance(cell, float) and np.isnan(cell)):
        return None, None
    s = str(cell).strip()
    if not s or s.lower() == "none":
        return None, None
    m = BAND_RE.match(s)
    if not m:
        raise ValueError(f"Unparseable Odds Band cell: {cell!r}")
    return float(m.group(1)), float(m.group(2))


def compute(raw_dir, slot, saved, lo, hi):
    df = pd.read_csv(raw_dir / f"slot{slot}_quals.tsv", sep="\t")
    df["Date"] = pd.to_datetime(df["Date"])
    since = pd.Timestamp(saved) + pd.Timedelta(days=1)
    sub = df[df["Date"] >= since].copy()

    if lo is not None:
        sub = sub[(sub["Odds_Exchange"] >= lo) & (sub["Odds_Exchange"] <= hi)]

    sub["is_win"] = sub["Position"].astype(str) == "1"
    bets = len(sub)
    wins = int(sub["is_win"].sum())
    winpct = round(100 * wins / bets, 2) if bets else 0
    pl_sp = np.where(sub["is_win"], sub["Odds_Numeric"], -1)
    pl_bf = np.where(sub["is_win"], (sub["Odds_Exchange"] - 1) * 0.95, -1)
    plsp_total = round(float(pl_sp.sum()), 2) if bets else 0
    plbf_total = round(float(pl_bf.sum()), 2) if bets else 0
    roisp = round(100 * plsp_total / bets, 2) if bets else 0
    roibf = round(100 * plbf_total / bets, 2) if bets else 0

    races = sub.drop_duplicates(subset=["Date", "RTime", "track"]).shape[0]
    winning_races = (sub[sub["is_win"]]
                      .drop_duplicates(subset=["Date", "RTime", "track"]).shape[0])
    racepct = round(100 * winning_races / races, 2) if races else 0

    sub["year"] = sub["Date"].dt.year
    by_year = (sub.assign(pl_bf=pl_bf).groupby("year")["pl_bf"]
               .sum().round(2).to_dict())

    return {
        "Bets": bets, "Wins": wins, "Win%": winpct, "P/L(SP)": plsp_total,
        "Races": races, "Race%": racepct, "ROI(SP)": roisp,
        "P/L(BF)": plbf_total, "ROI(BF)": roibf, "ByYear": by_year,
    }


def update_workbook(account, cfg):
    path = BASE / cfg["wb"]
    raw_dir = BASE / cfg["raw"]
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}
    year_cols = {h: col[h] for h in headers if h.startswith("P/L(BF) ") and h[8:].isdigit()}

    print(f"\n=== {account} ({cfg['wb']}) ===")
    changed = []
    for row in ws.iter_rows(min_row=2):
        slot = row[col["Slot"] - 1].value
        if slot not in cfg["slots"]:
            continue
        name = row[col["Name"] - 1].value
        saved = row[col["Saved"] - 1].value
        band_cell = row[col["Odds Band"] - 1].value
        lo, hi = parse_band(band_cell)

        old_bets = row[col["Bets"] - 1].value
        old_plbf = row[col["P/L(BF)"] - 1].value

        r = compute(raw_dir, slot, saved, lo, hi)

        row[col["Bets"] - 1].value = r["Bets"]
        row[col["Wins"] - 1].value = r["Wins"]
        row[col["Win%"] - 1].value = r["Win%"]
        row[col["P/L(SP)"] - 1].value = r["P/L(SP)"]
        row[col["Races"] - 1].value = r["Races"]
        row[col["Race%"] - 1].value = r["Race%"]
        row[col["ROI(SP)"] - 1].value = r["ROI(SP)"]
        row[col["P/L(BF)"] - 1].value = r["P/L(BF)"]
        row[col["ROI(BF)"] - 1].value = r["ROI(BF)"]

        # clear all existing year cells first, then fill in what this
        # system actually has - handles a year disappearing (shouldn't
        # happen since we only ever add history) as well as a new one.
        for h, c in year_cols.items():
            row[c - 1].value = None
        for y, val in r["ByYear"].items():
            h = f"P/L(BF) {y}"
            if h not in year_cols:
                raise ValueError(
                    f"{account} slot {slot}: year {y} has no column in "
                    f"the workbook - add it before re-running.")
            row[year_cols[h] - 1].value = val

        changed.append((slot, name, old_bets, r["Bets"], old_plbf, r["P/L(BF)"]))

    for slot, name, ob, nb, opl, npl in sorted(changed):
        print(f"  slot {slot:>3} {name[:40]:<40} "
              f"Bets {ob:>6} -> {nb:>6}   P/L(BF) {opl:>9.2f} -> {npl:>9.2f}")

    wb.save(path)
    print(f"  Saved {len(changed)} updated rows to {path.name}")
    missing = set(cfg["slots"]) - {c[0] for c in changed}
    if missing:
        print(f"  *** WARNING: slots not found in workbook: {sorted(missing)}")


def main():
    for account, cfg in TARGETS.items():
        update_workbook(account, cfg)


if __name__ == "__main__":
    main()
