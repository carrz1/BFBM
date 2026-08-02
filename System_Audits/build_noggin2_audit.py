import re
import pandas as pd
import numpy as np
from pathlib import Path

RAW = Path(r"E:/racing/AI/GitHub/BFBM/System_Audits/noggin2_raw")

# slot: (name, saved_date)
# Full list captured 2026-07-31 from v4savedsystems.php?userchoicego=1 (period=999,
# "Performance since saved date") for the noggin2 account. All 88 slots.
SLOTS = {
    1: ("MM AW FTO 2.01 - 40.00 AW", "2024-02-01"),
    2: ("SS MM AW FIRST RUN TRAINER 10.01 -30.00", "2024-02-01"),
    3: ("YORK EBOR 2023 14.01 - 1000.00", "2023-08-22"),
    4: ("NH HDGRdiff 10.01 - 400.00", "2024-08-26"),
    5: ("HDGRdiff FLAT AW 8.01 - 80.00", "2024-08-26"),
    6: ("SS D & G winners v1", "2024-03-03"),
    7: ("IRE HDGRdiff", "2024-10-07"),
    8: ("SS AW HDGRdiff TURF LTO 8.01 - 500", "2026-01-11"),
    9: ("RyanK HDGR FT", "2022-06-13"),
    10: ("EasterbyTD HDGR FTO", "2022-06-13"),
    11: ("HDGRdiff class6 - 7", "2024-10-07"),
    12: ("YORK 6f", "2024-02-18"),
    13: ("SS NH CHP LTO 2.01 - 300.00 IRE", "2023-01-28"),
    14: ("SS NH CHP LTO GB and FR 2.01 - 300.00", "2023-01-28"),
    15: ("ss CURRAGH HDGRdiff all BSPs", "2023-09-10"),
    16: ("FLAT CATT 5YOs allBSPs", "2026-05-21"),
    17: ("SS 5YO TT RHTRACKS 16.01-1000.00", "2026-05-21"),
    18: ("ss YORK HDGRdiff 7f-1m1f all BSPs", "2023-09-10"),
    19: ("ss YORK HDGRdiff 5f all BSPs", "2023-09-10"),
    20: ("ss YORK HDGRdiff 0 runs 12.01-1000.00", "2023-09-10"),
    21: ("HDGRdiff class6 - 7 BSP12.01-300", "2024-10-08"),
    22: ("FTO AW 14.01 - 150.00 CGs", "2024-01-31"),
    25: ("SS DUNDALK 5YOFMs HDGRdiff SS ALL BSPs", "2024-02-10"),
    26: ("SS DUNDALK 8YOs SS", "2024-02-10"),
    28: ("IRISH FLAT FMs 10.00-500.00", "2025-09-27"),
    29: ("SS AW NEWC STR 5YOs 5.51 - 65.00", "2024-02-10"),
    30: ("Gosden 1m - 1m6f", "2022-06-17"),
    31: ("D&G 7.01 - 300.00", "2024-09-11"),
    32: ("Evans PD", "2018-05-14"),
    33: ("EvansPD Chp Blnkers", "2018-05-14"),
    34: ("FaheyR FAVs", "2022-08-01"),
    35: ("GoldieJS 2021", "2021-05-19"),
    36: ("SS NEWC STR HDGRdiff 4.01 - 500.00", "2026-03-21"),
    37: ("GosdenJHM 4YOs 5YOs", "2018-05-14"),
    38: ("Gosden JHM Part 3", "2018-05-14"),
    39: ("Elliott Gordon Hurdles NH Flat ALL BSPs", "2022-02-06"),
    40: ("GoldieJS Ayr MAY JUNE 2021", "2021-05-19"),
    41: ("SS Ascot STR HDGR diff", "2022-10-15"),
    42: ("HaggasWJ", "2018-05-14"),
    43: ("SS HDGR 1-5 DSLR BSP 3.51-1000.00", "2024-10-08"),
    44: ("NH NAAS", "2024-02-10"),
    45: ("SS Chapstow 6YO", "2023-08-24"),
    47: ("NewmR ApplebyC", "2019-04-15"),
    48: ("SS NH NEWBURY", "2024-02-10"),
    49: ("SS NH CHP LTO 2.01 - 300.00", "2021-12-10"),
    51: ("SS NH 1 RUN 8.01 - 50.00", "2022-02-11"),
    52: ("SS AW HDGRdiff 7F - 1m.5F 14.01 - 200.00", "2026-02-21"),
    54: ("NewmR BeckettRM", "2019-04-15"),
    55: ("SS AW 1 RUN 5.01 - 100.00", "2022-02-10"),
    56: ("SS AW NEWC STR 4YOs 12.01 - 1000.00", "2024-02-10"),
    58: ("MidgleyPT 6f-7f", "2018-05-28"),
    61: ("IRE NH RH TRACKS 12.01 - 65.00", "2026-02-22"),
    62: ("ss IRE DSLRx2 14.01 - 1000.00", "2023-09-06"),
    63: ("ss THIRSK 7f", "2023-09-09"),
    64: ("SS AW 6f 6YO - 9YO all BSPs", "2025-08-19"),
    65: ("Newb - BinSuroor", "2019-04-11"),
    66: ("OBrienAP", "2026-05-30"),
    67: ("OKeef Jedd", "2018-05-28"),
    68: ("SS FLAT 1-7 DSLR 8YO", "2025-09-03"),
    69: ("HDGRdiff 5YO-8YO 10.5-65.00", "2025-08-15"),
    70: ("SS MM CLASS 1-3 ClaimJock 14.01 - 65.00", "2024-05-11"),
    71: ("SS BLNK FT ALL BSPs", "2025-02-05"),
    72: ("SS Salis 1 place at Track", "2023-07-15"),
    73: ("2021 HannonR Jr 4YO 2021", "2021-05-10"),
    74: ("SS NH HDGRdiff CHP LTO V today 2.01 - 100.00", "2026-04-09"),
    75: ("2021 MeehanB 2021", "2021-05-11"),
    76: ("SHORT TRAVELLERS 6.01 - 400.00", "2026-04-13"),
    77: ("MidgleyPT 5F 2021", "2021-05-13"),
    78: ("NH BF 7lbs LTO", "2021-02-01"),
    79: ("OBrienAP ROYAL ASCOT", "2021-06-16"),
    80: ("SmartB 2YOs", "2018-05-29"),
    81: ("SS NH DONCASTER SS 02.03.24", "2024-03-02"),
    83: ("SS HDGRdiff noHDGR 3.50 - 200.00", "2026-04-26"),
    84: ("MidgleyPT 5F 2021 see Tracks", "2021-07-09"),
    85: ("Bin Suroor 2YO TURF", "2018-05-29"),
    86: ("Bin Suroor 4YO 5YO", "2018-05-29"),
    88: ("SS NH Elliott Gordon Hurdles NH Flat ALL BSPs", "2026-04-29"),
    89: ("SS WINDOP FTO and WINDOP1RUN 10.01-1000.00", "2025-02-21"),
    90: ("NH JOCK & TR 1 DAY 12.01 - 80.00", "2025-05-21"),
    91: ("2021 HannonR Jr FESTIVALS FTOv1", "2021-09-10"),
    92: ("SS CHELMSFORD 4.01 - 100.00", "2024-02-10"),
    93: ("2012 Williams I 2021", "2021-06-15"),
    94: ("ss YARM HDGR Diff all BSPs", "2023-09-10"),
    95: ("SS OmearaD HDGRdiff", "2022-10-15"),
    96: ("YORK MAY MEETING 240 - 300 dslr", "2021-05-12"),
    97: ("YORK MAY MEETING 16-20 DSLR", "2021-05-12"),
    99: ("1 JOCK & 1 TR DAY 12.01 - 80.00", "2025-08-15"),
    100: ("SS FLAT 1-7 DSLR 7YO 6-16SOF", "2025-09-03"),
}

ODDS_RE = re.compile(r"(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)")

# Words glued directly onto the low side of a number range that mark it as
# something other than an odds band (e.g. "class6 - 7" = race class, not price).
PRECEDING_EXCLUDE = {"class"}
# Words immediately after a number range that mark it as something other than
# an odds band (e.g. "1-7 DSLR" = days since last run; "1-3 ClaimJock" = a
# claiming-class jockey category; "6-16SOF" = some other non-price category).
FOLLOWING_EXCLUDE = {"dslr", "dslrs", "sof", "claimjock"}


def parse_odds_band(name):
    """Find the first regex match in `name` that looks like a genuine odds
    band rather than an incidental number range (race class, DSLR count,
    etc). Returns (lo, hi, note) — note is set when a match was found but
    rejected, or when values looked invalid, so it can be surfaced for
    manual review rather than silently guessed."""
    notes = []
    for m in ODDS_RE.finditer(name):
        lo, hi = float(m.group(1)), float(m.group(2))

        preceding_word_match = re.search(r"([A-Za-z]+)$", name[: m.start()])
        preceding_word = preceding_word_match.group(1).lower() if preceding_word_match else ""
        following_word_match = re.match(r"\s*([A-Za-z]+)", name[m.end() :])
        following_word = following_word_match.group(1).lower() if following_word_match else ""

        if preceding_word in PRECEDING_EXCLUDE:
            notes.append(f"skipped '{m.group(0)}' (preceded by '{preceding_word}', not odds)")
            continue
        if following_word in FOLLOWING_EXCLUDE:
            notes.append(f"skipped '{m.group(0)}' (followed by '{following_word}', not odds)")
            continue
        if lo < 1.0 or hi < 1.0:
            notes.append(f"skipped '{m.group(0)}' (value <1.0, not odds)")
            continue

        return lo, hi, ("; ".join(notes) if notes else None)

    return None, None, ("; ".join(notes) if notes else None)


def compute(slot, name, saved_date):
    df = pd.read_csv(RAW / f"slot{slot}_quals.tsv", sep="\t")
    df["Date"] = pd.to_datetime(df["Date"])
    saved = pd.Timestamp(saved_date)
    since = saved + pd.Timedelta(days=1)
    sub = df[df["Date"] >= since].copy()

    lo, hi, note = parse_odds_band(name)
    if lo is not None:
        sub = sub[(sub["Odds_Exchange"] >= lo) & (sub["Odds_Exchange"] <= hi)]

    sub["is_win"] = sub["Position"].astype(str) == "1"
    bets = len(sub)
    wins = int(sub["is_win"].sum())
    winpct = round(100 * wins / bets, 2) if bets else 0
    pl_sp = np.where(sub["is_win"], sub["Odds_Numeric"], -1)
    pl_bf = np.where(sub["is_win"], (sub["Odds_Exchange"] - 1) * 0.95, -1)
    plsp_total = round(pl_sp.sum(), 2) if bets else 0
    plbf_total = round(pl_bf.sum(), 2) if bets else 0
    roisp = round(100 * plsp_total / bets, 2) if bets else 0
    roibf = round(100 * plbf_total / bets, 2) if bets else 0

    races = sub.drop_duplicates(subset=["Date", "RTime", "track"]).shape[0]
    winning_races = sub[sub["is_win"]].drop_duplicates(subset=["Date", "RTime", "track"]).shape[0]
    racepct = round(100 * winning_races / races, 2) if races else 0

    sub["year"] = sub["Date"].dt.year
    pl_bf_by_year = sub.assign(pl_bf=pl_bf).groupby("year")["pl_bf"].sum().round(2).to_dict()

    return {
        "Slot": slot,
        "Name": name,
        "Saved": saved_date,
        "Odds Band": f"{lo} - {hi}" if lo is not None else "none",
        "Odds Parsing Note": note or "",
        "Bets": bets,
        "Wins": wins,
        "Win%": winpct,
        "P/L(SP)": plsp_total,
        "Races": races,
        "Race%": racepct,
        "ROI(SP)": roisp,
        "P/L(BF)": plbf_total,
        "ROI(BF)": roibf,
        "P/L(BF)ByYear": pl_bf_by_year,
    }


if __name__ == "__main__":
    import openpyxl
    from openpyxl.utils import get_column_letter

    rows = []
    missing = []
    for slot, (name, saved) in SLOTS.items():
        if not (RAW / f"slot{slot}_quals.tsv").exists():
            missing.append(slot)
            continue
        rows.append(compute(slot, name, saved))

    if missing:
        print(f"Skipped {len(missing)} slots with no raw tsv yet: {missing}")

    years = sorted({y for r in rows for y in r["P/L(BF)ByYear"].keys()})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "noggin2"
    headers = [
        "Slot", "Name", "Saved", "Odds Band", "Odds Parsing Note", "Bets",
        "Wins", "Win%", "P/L(SP)", "Races", "Race%", "ROI(SP)", "P/L(BF)",
        "ROI(BF)",
    ] + [f"P/L(BF) {y}" for y in years]
    ws.append(headers)
    for r in rows:
        row = [r["Slot"], r["Name"], r["Saved"], r["Odds Band"],
               r["Odds Parsing Note"], r["Bets"], r["Wins"], r["Win%"],
               r["P/L(SP)"], r["Races"], r["Race%"], r["ROI(SP)"],
               r["P/L(BF)"], r["ROI(BF)"]]
        for y in years:
            row.append(r["P/L(BF)ByYear"].get(y, ""))
        ws.append(row)

    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

    out = Path(__file__).parent / "HRB_System_Performance_Audit_noggin2.xlsx"
    wb.save(out)
    print(f"Wrote {out} with {len(rows)} systems ({len(missing)} still missing raw data).")
