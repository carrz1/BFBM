"""noggin5 (account 5) system audit - built entirely from in-browser fetch +
JS-side compute (no raw TSV cached to disk), same pipeline as accounts 1/3/4.
Same odds-band-parse / saved+1-date-filter / BF-commission-formula pipeline as
build_noggin2_audit.py, using the account 4 exclusion word lists (class, week,
wk, n preceding; dslr, dslrs, sof, claimjock, run, runs, runtr following;
'-' or 'to' as range separator).

Mid-session bug found and fixed before any rows were finalized: the TSV export
values are double-quoted (e.g. "2" not 2), so the initial win-detection check
(`pos === '1'`) never matched and every early test computed 0 wins. Fixed by
stripping quotes from every column before comparison; all 68 slots below were
computed (or recomputed) after the fix.

A large number of slots on this account hit the 10,000-row download cap
(flagged in Note) - more than any prior account, likely because several of
these systems date back to 2018-2021 and have accumulated large qualifier
histories.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "HRB_System_Performance_Audit_noggin5.xlsx"

CAP_NOTE = "CAPPED AT 10,000-ROW DOWNLOAD LIMIT - true post-saved-date bet count may be higher, review manually"
VERIFIED_NOTE = "Hit the 10,000-row cap on the unfiltered download; re-verified via a server-side Odds(BFSP) pre-filter (see PROJECT.md) - the earliest row in that filtered download predates the saved date, confirming this figure is complete, not truncated."

ROWS = [
    {"Slot":1,"Name":"SS NH Diff HDGR Diff CODE","Saved":"2023-01-27","OddsBand":"none","Note":"","Bets":422,"Wins":38,"Win%":9.0,"P/L(SP)":-145.41,"Races":344,"Race%":11.05,"ROI(SP)":-34.46,"P/L(BF)":59.36,"ROI(BF)":14.07,"ByYear":{2023:-42.4115,2024:-80.0375,2025:199.178,2026:-17.3735}},
    {"Slot":2,"Name":"NH 2.01 - 500.00 5lbs LTO 2.01 to 125to1","Saved":"2021-02-16","OddsBand":"2.01 - 500","Note":"","Bets":1086,"Wins":83,"Win%":7.64,"P/L(SP)":-105.05,"Races":1028,"Race%":8.07,"ROI(SP)":-9.67,"P/L(BF)":202.43,"ROI(BF)":18.64,"ByYear":{2021:-31.4185,2022:149.9755,2023:33.997,2024:28.329,2025:25.6295,2026:-4.086}},
    {"Slot":3,"Name":"SS TT 5.01 - 100.00","Saved":"2023-01-31","OddsBand":"5.01 - 100","Note":VERIFIED_NOTE,"Bets":1423,"Wins":116,"Win%":8.15,"P/L(SP)":-210.68,"Races":1339,"Race%":8.66,"ROI(SP)":-14.81,"P/L(BF)":183.11,"ROI(BF)":12.87,"ByYear":{2023:86.7815,2024:-21.227,2025:38.119,2026:79.4395}},
    {"Slot":4,"Name":"FOR TO 1plus run 0 Runs TR","Saved":"2025-01-01","OddsBand":"none","Note":"","Bets":0,"Wins":0,"Win%":0,"P/L(SP)":0,"Races":0,"Race%":0,"ROI(SP)":0,"P/L(BF)":0,"ROI(BF)":0,"ByYear":{}},
    {"Slot":5,"Name":"SS 9YO 10YO CHASES 19plusSOF 40.01-1000.00","Saved":"2026-05-15","OddsBand":"40.01 - 1000","Note":"","Bets":2,"Wins":0,"Win%":0,"P/L(SP)":-2,"Races":2,"Race%":0,"ROI(SP)":-100,"P/L(BF)":-2,"ROI(BF)":-100,"ByYear":{2026:-2}},
    {"Slot":7,"Name":"SSwonatTrackLTO 8.01-18.00","Saved":"2025-01-30","OddsBand":"8.01 - 18","Note":VERIFIED_NOTE,"Bets":308,"Wins":26,"Win%":8.44,"P/L(SP)":-66,"Races":295,"Race%":8.81,"ROI(SP)":-21.43,"P/L(BF)":-18.32,"ROI(BF)":-5.95,"ByYear":{2025:4.3795,2026:-22.6975}},
    {"Slot":8,"Name":"TURF IRE FMs 14.01 - 80.00","Saved":"2025-02-04","OddsBand":"14.01 - 80","Note":VERIFIED_NOTE,"Bets":2939,"Wins":110,"Win%":3.74,"P/L(SP)":-918.5,"Races":2009,"Race%":5.48,"ROI(SP)":-31.25,"P/L(BF)":-214.13,"ROI(BF)":-7.29,"ByYear":{2025:40.432,2026:-254.557}},
    {"Slot":9,"Name":"FLAT & AW DandG 6F 20.00 - 80.00","Saved":"2025-12-18","OddsBand":"20 - 80","Note":VERIFIED_NOTE,"Bets":196,"Wins":5,"Win%":2.55,"P/L(SP)":-80,"Races":120,"Race%":4.17,"ROI(SP)":-40.82,"P/L(BF)":-27.81,"ROI(BF)":-14.19,"ByYear":{2025:-5,2026:-22.809}},
    {"Slot":10,"Name":"JClaim v LTO 6.01 - 150.00","Saved":"2023-01-19","OddsBand":"6.01 - 150","Note":VERIFIED_NOTE,"Bets":6123,"Wins":368,"Win%":6.01,"P/L(SP)":-1477.34,"Races":5138,"Race%":7.16,"ROI(SP)":-24.13,"P/L(BF)":88.18,"ROI(BF)":1.44,"ByYear":{2023:-93.932,2024:-52.89,2025:150.154,2026:84.852}},
    {"Slot":11,"Name":"C Placed NH GB IRE 7lb LTO Pt 1","Saved":"2021-02-21","OddsBand":"none","Note":"","Bets":79,"Wins":5,"Win%":6.33,"P/L(SP)":-36.42,"Races":79,"Race%":6.33,"ROI(SP)":-46.1,"P/L(BF)":-24.36,"ROI(BF)":-30.84,"ByYear":{2021:-18,2022:-1.6525,2023:-1.73,2024:16.55,2025:-11.53,2026:-8}},
    {"Slot":13,"Name":"AW Newc 6F 7F longest DSLR 9.01 - 200.00","Saved":"2026-01-11","OddsBand":"9.01 - 200","Note":"","Bets":44,"Wins":0,"Win%":0,"P/L(SP)":-44,"Races":32,"Race%":0,"ROI(SP)":-100,"P/L(BF)":-44,"ROI(BF)":-100,"ByYear":{2026:-44}},
    {"Slot":14,"Name":"SS MM 0 RUNS 14.00 - 200.00","Saved":"2025-08-29","OddsBand":"14 - 200","Note":"","Bets":82,"Wins":4,"Win%":4.88,"P/L(SP)":-19,"Races":71,"Race%":5.63,"ROI(SP)":-23.17,"P/L(BF)":-6.97,"ROI(BF)":-8.5,"ByYear":{2025:-2,2026:-4.9685}},
    {"Slot":15,"Name":"C Placed NH GB IRE 0lbs LTO","Saved":"2021-02-21","OddsBand":"none","Note":"","Bets":1231,"Wins":146,"Win%":11.86,"P/L(SP)":-39.19,"Races":1154,"Race%":12.65,"ROI(SP)":-3.18,"P/L(BF)":207.81,"ROI(BF)":16.88,"ByYear":{2021:-46.878,2022:119.998,2023:-8.41,2024:6.183,2025:162.211,2026:-25.2965}},
    {"Slot":16,"Name":"NH HDGRdiff 10.01 - 300.00","Saved":"2025-09-05","OddsBand":"10.01 - 300","Note":VERIFIED_NOTE,"Bets":643,"Wins":34,"Win%":5.29,"P/L(SP)":15,"Races":574,"Race%":5.92,"ROI(SP)":2.33,"P/L(BF)":246.25,"ROI(BF)":38.3,"ByYear":{2025:99.803,2026:146.444}},
    {"Slot":18,"Name":"CHELT 2022 HDGR 0 runs 14.01-65.00","Saved":"2022-03-14","OddsBand":"14.01 - 65","Note":"","Bets":51,"Wins":2,"Win%":3.92,"P/L(SP)":-13,"Races":33,"Race%":6.06,"ROI(SP)":-25.49,"P/L(BF)":-3.94,"ROI(BF)":-7.73,"ByYear":{2022:-10,2023:9.4415,2024:-11,2025:-6,2026:13.617}},
    {"Slot":21,"Name":"SS AW 7f 2yo3yo 8.01 - 500.00 v1","Saved":"2022-02-01","OddsBand":"8.01 - 500","Note":"","Bets":638,"Wins":30,"Win%":4.7,"P/L(SP)":-85.5,"Races":320,"Race%":9.38,"ROI(SP)":-13.4,"P/L(BF)":145.42,"ROI(BF)":22.79,"ByYear":{2022:28.923,2023:-124.8,2024:36.1725,2025:248.4975,2026:-43.3765}},
    {"Slot":22,"Name":"NH BottomWeights 10.00 - 400.00","Saved":"2026-01-12","OddsBand":"10 - 400","Note":VERIFIED_NOTE,"Bets":955,"Wins":45,"Win%":4.71,"P/L(SP)":-148.5,"Races":634,"Race%":7.1,"ROI(SP)":-15.55,"P/L(BF)":192.95,"ROI(BF)":20.2,"ByYear":{2026:192.95}},
    {"Slot":23,"Name":"SS 4YO 5F 5.5F 9.01 - 1000.00","Saved":"2022-06-26","OddsBand":"9.01 - 1000","Note":"","Bets":164,"Wins":18,"Win%":10.98,"P/L(SP)":97.5,"Races":147,"Race%":12.24,"ROI(SP)":59.45,"P/L(BF)":190.11,"ROI(BF)":115.92,"ByYear":{2022:41.0175,2023:14.635,2024:120.587,2025:22.8705,2026:-9}},
    {"Slot":26,"Name":"CHELT IRISH 2023","Saved":"2023-03-16","OddsBand":"none","Note":"","Bets":259,"Wins":12,"Win%":4.63,"P/L(SP)":-30.87,"Races":45,"Race%":26.67,"ROI(SP)":-11.92,"P/L(BF)":121.57,"ROI(BF)":46.94,"ByYear":{2023:-26,2024:-16.6125,2025:226.3675,2026:-62.1835}},
    {"Slot":27,"Name":"TURF HDGR LTO noHDGR 4.5 -500.00","Saved":"2025-03-16","OddsBand":"4.5 - 500","Note":VERIFIED_NOTE,"Bets":1220,"Wins":108,"Win%":8.85,"P/L(SP)":50.39,"Races":1087,"Race%":9.94,"ROI(SP)":4.13,"P/L(BF)":514.89,"ROI(BF)":42.2,"ByYear":{2025:529.3025,2026:-14.4085}},
    {"Slot":28,"Name":"Chepstow 7YOs","Saved":"2018-07-03","OddsBand":"none","Note":"","Bets":300,"Wins":42,"Win%":14.0,"P/L(SP)":38.42,"Races":202,"Race%":20.79,"ROI(SP)":12.81,"P/L(BF)":93.89,"ROI(BF)":31.3,"ByYear":{2018:18.3935,2019:4.374,2020:-2.6385,2021:52.7205,2022:6.134,2023:37.1765,2024:-0.3715,2025:-19.599,2026:-2.3}},
    {"Slot":30,"Name":"Ascot HUNT CUP AW WIN LTO","Saved":"2018-06-20","OddsBand":"none","Note":"","Bets":32,"Wins":2,"Win%":6.25,"P/L(SP)":0,"Races":12,"Race%":16.67,"ROI(SP)":0,"P/L(BF)":0.88,"ROI(BF)":2.73,"ByYear":{2019:-3,2020:7.825,2021:17.05,2022:-1,2023:-1,2024:-7,2025:-4,2026:-8}},
    {"Slot":31,"Name":"2021 MeehanBJ  TURF 2021","Saved":"2021-05-07","OddsBand":"none","Note":"","Bets":63,"Wins":6,"Win%":9.52,"P/L(SP)":46.13,"Races":63,"Race%":9.52,"ROI(SP)":73.22,"P/L(BF)":83.23,"ROI(BF)":132.11,"ByYear":{2021:-10,2022:-0.2745,2023:-1.0225,2024:112.8165,2025:-12.29,2026:-6}},
    {"Slot":32,"Name":"SS CHEP 5F 6F all BSPs","Saved":"2026-05-15","OddsBand":"none","Note":"","Bets":36,"Wins":6,"Win%":16.67,"P/L(SP)":22.71,"Races":13,"Race%":46.15,"ROI(SP)":63.08,"P/L(BF)":32.05,"ROI(BF)":89.04,"ByYear":{2026:32.054}},
    {"Slot":33,"Name":"Donc Rnd Course 9YO 10YO","Saved":"2019-04-01","OddsBand":"none","Note":"","Bets":37,"Wins":3,"Win%":8.11,"P/L(SP)":-5,"Races":31,"Race%":9.68,"ROI(SP)":-13.51,"P/L(BF)":2.95,"ROI(BF)":7.99,"ByYear":{2019:-10,2021:11.345,2022:-3,2023:12.39,2024:3.22,2025:-9,2026:-2}},
    {"Slot":34,"Name":"Bev EllisonB","Saved":"2019-04-17","OddsBand":"none","Note":"","Bets":108,"Wins":5,"Win%":4.63,"P/L(SP)":46.91,"Races":100,"Race%":5.0,"ROI(SP)":43.44,"P/L(BF)":112.44,"ROI(BF)":104.11,"ByYear":{2019:-8,2020:-7,2021:-10,2022:-5.692,2023:-12,2024:191.903,2025:-21.77,2026:-15}},
    {"Slot":35,"Name":"Ascot 7f","Saved":"2018-06-20","OddsBand":"none","Note":"","Bets":83,"Wins":7,"Win%":8.43,"P/L(SP)":43.5,"Races":20,"Race%":35.0,"ROI(SP)":52.41,"P/L(BF)":97.68,"ROI(BF)":117.69,"ByYear":{2018:-10,2019:-6,2020:-2,2021:26.751,2022:5.2235,2023:98.2,2024:-8,2025:-4.68,2026:-1.8155}},
    {"Slot":36,"Name":"SS 2YOs CHEP STRT all SPs","Saved":"2026-05-15","OddsBand":"none","Note":"","Bets":23,"Wins":2,"Win%":8.7,"P/L(SP)":-9,"Races":8,"Race%":25.0,"ROI(SP)":-39.13,"P/L(BF)":-6.13,"ROI(BF)":-26.66,"ByYear":{2026:-6.1325}},
    {"Slot":38,"Name":"SS HDGR diff 1-5DSLR all BSPs","Saved":"2023-01-04","OddsBand":"none","Note":"skipped '1-5' (followed by 'dslr', not odds)","Bets":86,"Wins":13,"Win%":15.12,"P/L(SP)":0.53,"Races":83,"Race%":15.66,"ROI(SP)":0.62,"P/L(BF)":-2.37,"ROI(BF)":-2.75,"ByYear":{2023:14.577,2024:-9.762,2025:-3.8875,2026:-3.295}},
    {"Slot":39,"Name":"HDGR DIFF v2 14.01 - 80.00","Saved":"2021-04-02","OddsBand":"14.01 - 80","Note":"CORRECTED 2026-08-01: original figure was truncated (the odds-filtered download alone still hits the 10,000-row cap without reaching the saved date). Fixed by pulling a second, date-restricted download (2021-2023) via a server-side Date(Year) filter and merging the two non-overlapping downloads. See PROJECT.md / CHANGELOG.md for the full writeup.","Bets":16928,"Wins":687,"Win%":4.06,"P/L(SP)":-3552.0,"Races":13913,"Race%":4.94,"ROI(SP)":-20.98,"P/L(BF)":419.14,"ROI(BF)":2.48,"ByYear":{2021:-95.991,2022:24.8525,2023:184.43,2024:128.2525,2025:-120.0965,2026:297.693}},
    {"Slot":40,"Name":"NH HDGRdiff 6YO 10.01-400.00","Saved":"2025-08-13","OddsBand":"10.01 - 400","Note":VERIFIED_NOTE,"Bets":343,"Wins":18,"Win%":5.25,"P/L(SP)":-89.5,"Races":317,"Race%":5.68,"ROI(SP)":-26.09,"P/L(BF)":-41.77,"ROI(BF)":-12.18,"ByYear":{2025:-2.411,2026:-39.356}},
    {"Slot":41,"Name":"HDGR DIFF v3 9.01 - 200.00","Saved":"2026-05-15","OddsBand":"9.01 - 200","Note":VERIFIED_NOTE,"Bets":1042,"Wins":54,"Win%":5.18,"P/L(SP)":-162.5,"Races":799,"Race%":6.76,"ROI(SP)":-15.6,"P/L(BF)":107.25,"ROI(BF)":10.29,"ByYear":{2026:107.255}},
    {"Slot":42,"Name":"HDGR DIFF v4 NH 3.01-300.00","Saved":"2026-05-15","OddsBand":"3.01 - 300","Note":VERIFIED_NOTE,"Bets":384,"Wins":28,"Win%":7.29,"P/L(SP)":-137.42,"Races":290,"Race%":9.66,"ROI(SP)":-35.79,"P/L(BF)":-81.37,"ROI(BF)":-21.19,"ByYear":{2026:-81.374}},
    {"Slot":43,"Name":"SS Chest BaldingAM 2023","Saved":"2023-08-06","OddsBand":"none","Note":"","Bets":73,"Wins":16,"Win%":21.92,"P/L(SP)":4.97,"Races":72,"Race%":22.22,"ROI(SP)":6.81,"P/L(BF)":10.79,"ROI(BF)":14.78,"ByYear":{2023:-2,2024:8.7665,2025:15.141,2026:-11.1155}},
    {"Slot":44,"Name":"SS 4.50 - 65.00 1-2 CODE RUNS","Saved":"2025-08-30","OddsBand":"4.5 - 65","Note":"","Bets":322,"Wins":32,"Win%":9.94,"P/L(SP)":-3.17,"Races":284,"Race%":11.27,"ROI(SP)":-0.98,"P/L(BF)":98.83,"ROI(BF)":30.69,"ByYear":{2025:89.927,2026:8.8985}},
    {"Slot":45,"Name":"Leicester DRAW","Saved":"2017-05-29","OddsBand":"none","Note":"","Bets":2289,"Wins":297,"Win%":12.98,"P/L(SP)":-252.18,"Races":1161,"Race%":25.58,"ROI(SP)":-11.02,"P/L(BF)":326.92,"ROI(BF)":14.28,"ByYear":{2017:-43.1335,2018:38.0365,2019:287.2365,2020:2.3445,2021:13.331,2022:76.141,2023:58.2035,2024:-74.1285,2025:9.9025,2026:-41.012}},
    {"Slot":49,"Name":"SS NH IRISH HDGRdiff CHP  7.51 plus","Saved":"2026-03-04","OddsBand":"none","Note":"","Bets":41,"Wins":3,"Win%":7.32,"P/L(SP)":-22.5,"Races":34,"Race%":8.82,"ROI(SP)":-54.88,"P/L(BF)":-20.43,"ROI(BF)":-49.84,"ByYear":{2026:-20.4345}},
    {"Slot":51,"Name":"SS Aug Sep 2yo3yo4yo pt1 6.01 - 1000.00","Saved":"2022-07-31","OddsBand":"6.01 - 1000","Note":"","Bets":340,"Wins":27,"Win%":7.94,"P/L(SP)":-38.5,"Races":242,"Race%":11.16,"ROI(SP)":-11.32,"P/L(BF)":28.37,"ROI(BF)":8.34,"ByYear":{2022:-33.179,2023:-68.3015,2024:137.122,2025:-7.2685}},
    {"Slot":52,"Name":"SS Aug Sep 2yo3yo IRE pt2 all BSPs","Saved":"2022-07-31","OddsBand":"none","Note":"","Bets":12,"Wins":0,"Win%":0,"P/L(SP)":-12,"Races":11,"Race%":0,"ROI(SP)":-100,"P/L(BF)":-12,"ROI(BF)":-100,"ByYear":{2022:-5,2023:-3,2024:-4}},
    {"Slot":53,"Name":"0 Runs TR  all BSPs","Saved":"2026-05-08","OddsBand":"none","Note":"Hit the 10,000-row cap on the unfiltered download (no odds band to pre-filter on); re-checked the raw download directly - the earliest row predates the saved date, confirming this figure is complete, not truncated.","Bets":211,"Wins":25,"Win%":11.85,"P/L(SP)":27.2,"Races":195,"Race%":12.82,"ROI(SP)":12.89,"P/L(BF)":68.37,"ROI(BF)":32.4,"ByYear":{2026:68.372}},
    {"Slot":54,"Name":"DRAW Newb 7f","Saved":"2017-05-28","OddsBand":"none","Note":"","Bets":91,"Wins":7,"Win%":7.69,"P/L(SP)":-5.5,"Races":27,"Race%":25.93,"ROI(SP)":-6.04,"P/L(BF)":21.24,"ROI(BF)":23.34,"ByYear":{2017:7.8415,2018:-3.911,2019:-3.975,2020:0.7515,2021:-13,2022:-10,2023:-1,2024:-8,2025:47.3,2026:5.234}},
    {"Slot":55,"Name":"SS 5F HDGRdiff 8.01-1000.00","Saved":"2026-04-07","OddsBand":"8.01 - 1000","Note":"","Bets":65,"Wins":4,"Win%":6.15,"P/L(SP)":18,"Races":49,"Race%":8.16,"ROI(SP)":27.69,"P/L(BF)":40.02,"ROI(BF)":61.57,"ByYear":{2026:40.023}},
    {"Slot":56,"Name":"SS Ascot STR","Saved":"2026-05-07","OddsBand":"none","Note":"","Bets":90,"Wins":8,"Win%":8.89,"P/L(SP)":-36.24,"Races":18,"Race%":44.44,"ROI(SP)":-40.27,"P/L(BF)":-29.3,"ROI(BF)":-32.56,"ByYear":{2026:-29.3035}},
    {"Slot":57,"Name":"SS FLAT NewTRNR 0-6 RUNS 3.01 - 65.00","Saved":"2025-10-12","OddsBand":"3.01 - 65","Note":"skipped '0-6' (followed by 'runs', not odds)","Bets":270,"Wins":32,"Win%":11.85,"P/L(SP)":58.96,"Races":240,"Race%":13.33,"ROI(SP)":21.84,"P/L(BF)":124.49,"ROI(BF)":46.11,"ByYear":{2025:-23,2026:147.4915}},
    {"Slot":59,"Name":"Muss BurkeKR","Saved":"2019-04-20","OddsBand":"none","Note":"","Bets":64,"Wins":10,"Win%":15.63,"P/L(SP)":0.09,"Races":58,"Race%":17.24,"ROI(SP)":0.14,"P/L(BF)":9.3,"ROI(BF)":14.53,"ByYear":{2019:-2,2020:-3,2021:3.2,2022:-6,2023:7.9265,2024:15.6885,2025:-5.0765,2026:-1.44}},
    {"Slot":60,"Name":"ss 9YO 10YO 11YOs  NH","Saved":"2022-12-05","OddsBand":"none","Note":"","Bets":326,"Wins":10,"Win%":3.07,"P/L(SP)":-142.75,"Races":134,"Race%":7.46,"ROI(SP)":-43.79,"P/L(BF)":71.09,"ROI(BF)":21.81,"ByYear":{2023:-54.11,2024:-49.8545,2025:35.721,2026:139.3305}},
    {"Slot":61,"Name":"2018 9YOs 10YOs 13.00 - 51.00 9YOs 10YOs","Saved":"2018-03-14","OddsBand":"13 - 51","Note":"","Bets":226,"Wins":7,"Win%":3.1,"P/L(SP)":-83,"Races":111,"Race%":6.31,"ROI(SP)":-36.73,"P/L(BF)":-12.54,"ROI(BF)":-5.55,"ByYear":{2018:66.0965,2019:13.55,2020:-18,2021:-2.1,2022:-19,2023:-30,2024:-23,2025:-3.7,2026:3.617}},
    {"Slot":62,"Name":"9YOs 10YOs 9YOs 10YOs","Saved":"2021-08-30","OddsBand":"none","Note":"","Bets":189,"Wins":5,"Win%":2.65,"P/L(SP)":-8.62,"Races":72,"Race%":6.94,"ROI(SP)":-4.56,"P/L(BF)":129.87,"ROI(BF)":68.71,"ByYear":{2021:-4,2022:26.5305,2023:57.683,2024:-32,2025:107.776,2026:-26.119}},
    {"Slot":63,"Name":"SS AW 7f 2yo3yo 8.01 - 500.00","Saved":"2022-01-18","OddsBand":"8.01 - 500","Note":"","Bets":777,"Wins":38,"Win%":4.89,"P/L(SP)":-85.5,"Races":366,"Race%":10.38,"ROI(SP)":-11.0,"P/L(BF)":180.38,"ROI(BF)":23.22,"ByYear":{2022:2.435,2023:-137.3555,2024:124.56,2025:235.215,2026:-44.473}},
    {"Slot":64,"Name":"SS NH 2018 15 CAREER RUNS v1","Saved":"2023-11-19","OddsBand":"none","Note":"","Bets":245,"Wins":23,"Win%":9.39,"P/L(SP)":6.75,"Races":224,"Race%":10.27,"ROI(SP)":2.76,"P/L(BF)":68.72,"ROI(BF)":28.05,"ByYear":{2023:-5,2024:14.369,2025:2.2015,2026:57.1485}},
    {"Slot":68,"Name":"Muss KirbyP","Saved":"2019-04-20","OddsBand":"none","Note":"","Bets":49,"Wins":9,"Win%":18.37,"P/L(SP)":6.09,"Races":48,"Race%":18.75,"ROI(SP)":12.43,"P/L(BF)":14.3,"ROI(BF)":29.19,"ByYear":{2019:-4,2020:-1.126,2021:-6,2022:-3.012,2023:-8.252,2024:1.878,2025:35.7905,2026:-0.9765}},
    {"Slot":71,"Name":"Chest BaldingAM","Saved":"2019-05-10","OddsBand":"none","Note":"","Bets":314,"Wins":65,"Win%":20.7,"P/L(SP)":25.06,"Races":300,"Race%":21.67,"ROI(SP)":7.98,"P/L(BF)":63.69,"ROI(BF)":20.28,"ByYear":{2019:-9.8315,2020:-6.025,2021:10.7195,2022:9.783,2023:24.558,2024:9.6975,2025:32.3335,2026:-7.5425}},
    {"Slot":72,"Name":"Chest Ryan KA","Saved":"2019-05-10","OddsBand":"none","Note":"","Bets":35,"Wins":5,"Win%":14.29,"P/L(SP)":0.5,"Races":34,"Race%":14.71,"ROI(SP)":1.43,"P/L(BF)":7.51,"ROI(BF)":21.45,"ByYear":{2019:-0.4705,2020:-6,2021:4.367,2022:-8,2023:-1,2024:-3,2025:24.6095,2026:-3}},
    {"Slot":73,"Name":"Chest WilliamsI","Saved":"2019-05-10","OddsBand":"none","Note":"","Bets":178,"Wins":14,"Win%":7.87,"P/L(SP)":-31.25,"Races":131,"Race%":10.69,"ROI(SP)":-17.56,"P/L(BF)":43.38,"ROI(BF)":24.37,"ByYear":{2019:-1.117,2020:-0.8445,2021:-24.837,2022:-5.639,2023:62.8,2024:-30,2025:33.5955,2026:9.4175}},
    {"Slot":74,"Name":"Ascot STR 5f - 7f 4YO","Saved":"2018-06-27","OddsBand":"none","Note":"","Bets":180,"Wins":18,"Win%":10.0,"P/L(SP)":-5.5,"Races":108,"Race%":16.67,"ROI(SP)":-3.06,"P/L(BF)":69.04,"ROI(BF)":38.36,"ByYear":{2018:-6.6985,2019:-11.3475,2020:26.7885,2021:-21.858,2022:76.292,2023:-8.625,2024:-5.674,2025:30.1625,2026:-10}},
    {"Slot":75,"Name":"SS JARDINE I MUSSELBURGH","Saved":"2025-04-13","OddsBand":"none","Note":"","Bets":46,"Wins":4,"Win%":8.7,"P/L(SP)":-3.75,"Races":43,"Race%":9.3,"ROI(SP)":-8.15,"P/L(BF)":13.93,"ROI(BF)":30.27,"ByYear":{2025:-10.7235,2026:24.65}},
    {"Slot":79,"Name":"Hamilton SPRINTS","Saved":"2018-07-03","OddsBand":"none","Note":"","Bets":136,"Wins":16,"Win%":11.76,"P/L(SP)":14.63,"Races":76,"Race%":21.05,"ROI(SP)":10.76,"P/L(BF)":40.63,"ROI(BF)":29.87,"ByYear":{2018:11.4365,2019:-13.702,2020:-7.1975,2021:0.846,2022:44.6555,2023:0.941,2024:13.5215,2025:-13,2026:3.125}},
    {"Slot":82,"Name":"Hamil 6f","Saved":"2018-07-03","OddsBand":"none","Note":"","Bets":106,"Wins":13,"Win%":12.26,"P/L(SP)":-7.8,"Races":94,"Race%":13.83,"ROI(SP)":-7.36,"P/L(BF)":6.79,"ROI(BF)":6.4,"ByYear":{2018:-9,2019:4.019,2020:-3.63,2021:-16,2022:10.6555,2023:3.2,2024:7.85,2025:7.3145,2026:2.379}},
    {"Slot":87,"Name":"Leicester DRAW v1A","Saved":"2021-09-20","OddsBand":"none","Note":"","Bets":322,"Wins":39,"Win%":12.11,"P/L(SP)":-88.71,"Races":205,"Race%":19.02,"ROI(SP)":-27.55,"P/L(BF)":-70.34,"ROI(BF)":-21.85,"ByYear":{2021:1.668,2022:-28.475,2023:-35.295,2024:-13.684,2025:14.521,2026:-9.0775}},
    {"Slot":88,"Name":"2021 9YOs 10YOs all BSPs","Saved":"2021-10-08","OddsBand":"none","Note":"","Bets":145,"Wins":5,"Win%":3.45,"P/L(SP)":-30.62,"Races":46,"Race%":10.87,"ROI(SP)":-21.12,"P/L(BF)":42.31,"ROI(BF)":29.18,"ByYear":{2022:-23.328,2023:76.683,2024:-31,2025:24.4615,2026:-4.502}},
    {"Slot":90,"Name":"NEW TR FT 3.50 - 500.00","Saved":"2025-08-25","OddsBand":"3.5 - 500","Note":VERIFIED_NOTE,"Bets":1341,"Wins":106,"Win%":7.9,"P/L(SP)":-291.12,"Races":1145,"Race%":9.26,"ROI(SP)":-21.71,"P/L(BF)":33.91,"ROI(BF)":2.53,"ByYear":{2025:55.045,2026:-21.13}},
    {"Slot":91,"Name":"NH FMs 61-240DSLR 30.00-300.00","Saved":"2025-08-24","OddsBand":"30 - 300","Note":"skipped '61-240' (followed by 'dslr', not odds); " + VERIFIED_NOTE,"Bets":316,"Wins":3,"Win%":0.95,"P/L(SP)":-137,"Races":289,"Race%":1.04,"ROI(SP)":-43.35,"P/L(BF)":25.37,"ROI(BF)":8.03,"ByYear":{2025:95.251,2026:-69.88}},
    {"Slot":92,"Name":"Red Scott Dixon","Saved":"2019-04-22","OddsBand":"none","Note":"","Bets":17,"Wins":1,"Win%":5.88,"P/L(SP)":0,"Races":16,"Race%":6.25,"ROI(SP)":0,"P/L(BF)":3.95,"ROI(BF)":23.24,"ByYear":{2019:-2,2020:-2,2021:14.95,2022:-4,2023:-3}},
    {"Slot":94,"Name":"Rip OmearaD","Saved":"2019-04-27","OddsBand":"none","Note":"","Bets":241,"Wins":34,"Win%":14.11,"P/L(SP)":-44.9,"Races":215,"Race%":15.81,"ROI(SP)":-18.63,"P/L(BF)":-19.66,"ROI(BF)":-8.16,"ByYear":{2019:-17.158,2020:1.0895,2021:-3.9275,2022:5.358,2023:-20.1695,2024:-1.92,2025:13.846,2026:3.2215}},
    {"Slot":97,"Name":"SS 4YO GB STRAIGHTS UK","Saved":"2022-06-07","OddsBand":"none","Note":"","Bets":224,"Wins":12,"Win%":5.36,"P/L(SP)":14.5,"Races":209,"Race%":5.74,"ROI(SP)":6.47,"P/L(BF)":249.01,"ROI(BF)":111.16,"ByYear":{2022:-20.686,2023:-22.12,2024:331.749,2025:-30.465,2026:-9.4715}},
    {"Slot":98,"Name":"Ascot HUNT CUP Hdgr","Saved":"2022-06-15","OddsBand":"none","Note":"","Bets":20,"Wins":0,"Win%":0,"P/L(SP)":-20,"Races":7,"Race%":0,"ROI(SP)":-100,"P/L(BF)":-20,"ROI(BF)":-100,"ByYear":{2023:-1,2024:-7,2025:-4,2026:-8}},
    {"Slot":99,"Name":"Hamilton SPRINTS ALL AGES","Saved":"2024-06-06","OddsBand":"none","Note":"","Bets":46,"Wins":8,"Win%":17.39,"P/L(SP)":82.21,"Races":27,"Race%":29.63,"ROI(SP)":178.72,"P/L(BF)":206.47,"ROI(BF)":448.85,"ByYear":{2024:32.27,2025:-4.089,2026:178.292}},
    {"Slot":100,"Name":"SS HDGR none BLKRS LTO all BSPs","Saved":"2022-06-18","OddsBand":"none","Note":"","Bets":403,"Wins":30,"Win%":7.44,"P/L(SP)":-45.57,"Races":395,"Race%":7.59,"ROI(SP)":-11.31,"P/L(BF)":88.45,"ROI(BF)":21.95,"ByYear":{2022:57.6875,2023:0.843,2024:7.6525,2025:-63.9395,2026:86.2095}},
]

ROWS.sort(key=lambda r: r["Slot"])
years = sorted({y for r in ROWS for y in r["ByYear"].keys()})

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "noggin5"
headers = ["Slot", "Name", "Saved", "Odds Band", "Odds Parsing Note", "Bets",
           "Wins", "Win%", "P/L(SP)", "Races", "Race%", "ROI(SP)", "P/L(BF)",
           "ROI(BF)"] + [f"P/L(BF) {y}" for y in years]
ws.append(headers)
for r in ROWS:
    row = [r["Slot"], r["Name"], r["Saved"], r["OddsBand"], r["Note"],
           r["Bets"], r["Wins"], r["Win%"], r["P/L(SP)"], r["Races"],
           r["Race%"], r["ROI(SP)"], r["P/L(BF)"], r["ROI(BF)"]]
    for y in years:
        row.append(r["ByYear"].get(y, ""))
    ws.append(row)

for i, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

wb.save(OUT)
print(f"Wrote {OUT} with {len(ROWS)} systems.")
